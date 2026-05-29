"""Excel template writer for competition result files.

Reads template .xlsx files, identifies grey-filled cells, writes computed solutions,
and generates validation reports. Always preserves original templates.
"""

from __future__ import annotations

import json
import logging
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from .models import ExperimentResult, Material, Pattern, Piece


def _is_grey_fill(fill: PatternFill) -> bool:
    """Check if a cell fill is grey (intended for data entry)."""
    if fill is None or fill.fgColor is None:
        return False
    try:
        fg = fill.fgColor
        if fg.type == "rgb" and fg.rgb:
            rgb = fg.rgb
            if len(rgb) == 8:
                rgb = rgb[2:]
            r, g, b = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
            if abs(r - g) < 30 and abs(g - b) < 30 and abs(r - b) < 30:
                if 100 < r < 240:
                    return True
        elif fg.type == "indexed":
            if fg.indexed in [15, 16, 48]:
                return True
    except Exception:
        pass
    return False


def _get_grey_cells(ws) -> list[dict[str, Any]]:
    """Identify all grey-filled cells in a worksheet."""
    grey_cells = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if _is_grey_fill(cell.fill):
                grey_cells.append({
                    "row": cell.row,
                    "col": cell.column,
                    "col_letter": get_column_letter(cell.column),
                    "coordinate": cell.coordinate,
                    "value": cell.value,
                })
    return grey_cells


def _find_adjacent_label(ws, row: int, col: int) -> str:
    """Try to find a label adjacent to a grey cell (look left and up)."""
    for c in range(col - 1, max(0, col - 4), -1):
        val = ws.cell(row=row, column=c).value
        if val is not None:
            return str(val).strip()
    for r in range(row - 1, max(0, row - 4), -1):
        val = ws.cell(row=r, column=col).value
        if val is not None:
            return str(val).strip()
    return ""


def _get_column_header(ws, col: int) -> str:
    """Get the column header."""
    for r in range(1, min(5, ws.max_row + 1)):
        val = ws.cell(row=r, column=col).value
        if val is not None:
            return str(val).strip()
    return ""


def _determine_cell_value(
    gc: dict[str, Any],
    result: ExperimentResult,
    materials: dict[str, Material],
    pieces: dict[str, Piece],
    ws: Any,
) -> Any:
    """Try to determine the appropriate value for a grey cell based on context."""
    label = gc.get("adjacent_label", "")
    col_header = gc.get("column_header", "")
    context = (label + " " + col_header).lower()

    s = result.solution
    pattern_map = {p.pattern_id: p for p in result.patterns}

    # Material context
    for mat_name in ["L01", "L02", "L03"]:
        if mat_name.lower() in context:
            mat = materials.get(mat_name)
            if mat is None:
                continue
            used_vol = sum(
                count * pattern_map[pid].used_volume
                for pid, count in s.pattern_usage.items()
                if pid in pattern_map and pattern_map[pid].material_name == mat_name
            )
            total_vol = mat.volume * mat.count
            if any(w in context for w in ["利用率", "utilization", "util"]):
                return f"{used_vol / total_vol * 100:.2f}%" if total_vol > 0 else "0%"
            if any(w in context for w in ["废料", "waste"]):
                return total_vol - used_vol
            if any(w in context for w in ["用量", "使用", "count", "数量"]):
                return sum(
                    count for pid, count in s.pattern_usage.items()
                    if pid in pattern_map and pattern_map[pid].material_name == mat_name
                )
            return f"{used_vol / total_vol * 100:.2f}%" if total_vol > 0 else "0%"

    # Piece context
    for pname in ["J01", "J02", "J03", "J04", "J05", "J06", "J07"]:
        if pname.lower() in context:
            if any(w in context for w in ["数量", "产量", "count", "quantity"]):
                return s.piece_counts.get(pname, 0)
            if any(w in context for w in ["收益", "profit"]):
                count = s.piece_counts.get(pname, 0)
                return count * pieces[pname].profit if pname in pieces else 0
            return s.piece_counts.get(pname, 0)

    # Overall metrics
    if any(w in context for w in ["总利用率", "overall util", "整体利用率"]):
        return f"{s.material_utilization * 100:.2f}%"
    if any(w in context for w in ["总废料", "total waste"]):
        return s.total_waste_volume
    if any(w in context for w in ["总收益", "total profit"]):
        return f"{s.total_profit:,.0f}" if s.total_profit > 1e6 else s.total_profit
    if any(w in context for w in ["总使用体积", "used volume"]):
        return s.total_used_volume

    return None


def _add_computed_solution_sheet(
    wb: openpyxl.Workbook,
    result: ExperimentResult,
    materials: dict[str, Material],
    pieces: dict[str, Piece],
    timing: dict[str, float] | None = None,
) -> None:
    """Add a comprehensive computed_solution sheet to the workbook."""
    sheet_name = "computed_solution"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=14, color="1F4E79")
    sub_font = Font(bold=True, size=12, color="2E75B6")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    s = result.solution
    row = 1

    ws.cell(row=row, column=1, value=f"Computed Solution - {result.problem_name.upper()}").font = title_font
    row += 2

    # Basic info
    ws.cell(row=row, column=1, value="Basic Information").font = sub_font
    row += 1
    info_items = [
        ("Status", s.status.value),
        ("Objective Value", f"{s.objective_value:.2f}"),
        ("Total Profit", f"{s.total_profit:,.2f}"),
        ("Total Used Volume", f"{s.total_used_volume:,}"),
        ("Total Waste Volume", f"{s.total_waste_volume:,}"),
        ("Total Material Volume", f"{s.total_material_volume:,}"),
        ("Material Utilization", f"{s.material_utilization * 100:.2f}%"),
        ("Waste Gap Ratio", f"{s.total_waste_volume / s.total_material_volume * 100:.4f}%" if s.total_material_volume > 0 else "N/A"),
        ("Utilization Gap to 100%", f"{(1.0 - s.material_utilization) * 100:.4f}%"),
        ("Upper Bound (utilization)", "1.0 (100%)" if "problem1" in result.problem_name else f"{s.upper_bound:,.2f}"),
        ("Lower Bound", f"{s.lower_bound:,.4f}" if s.lower_bound < 2.0 else f"{s.lower_bound:,.2f}"),
        ("Gap", f"{s.gap * 100:.4f}%"),
        ("Solve Time (s)", f"{s.solve_time_seconds:.2f}"),
        ("Optimality Note", "Pattern-library optimal (not necessarily global optimal without complete pattern enumeration)"),
    ]
    for label, value in info_items:
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = Font(bold=True)
        c1.border = thin_border
        c2 = ws.cell(row=row, column=2, value=value)
        c2.border = thin_border
        row += 1
    row += 1

    # Timing breakdown
    if timing:
        ws.cell(row=row, column=1, value="Runtime Breakdown").font = sub_font
        row += 1
        for key, val in timing.items():
            ws.cell(row=row, column=1, value=key).font = Font(bold=True)
            ws.cell(row=row, column=1).border = thin_border
            ws.cell(row=row, column=2, value=f"{val:.2f}s").border = thin_border
            row += 1
        row += 1

    # Piece counts
    ws.cell(row=row, column=1, value="Piece Production Quantities").font = sub_font
    row += 1
    piece_headers = ["Piece", "Quantity", "Volume/Unit", "Total Volume", "Unit Profit", "Total Profit"]
    for j, h in enumerate(piece_headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = header_font; c.fill = header_fill; c.border = thin_border
    row += 1
    for pname in sorted(pieces.keys()):
        count = s.piece_counts.get(pname, 0)
        piece = pieces[pname]
        vals = [pname, count, piece.volume, count * piece.volume, piece.profit, count * piece.profit]
        for j, v in enumerate(vals, 1):
            ws.cell(row=row, column=j, value=v).border = thin_border
        row += 1
    row += 1

    # Material utilization breakdown
    ws.cell(row=row, column=1, value="Material Utilization Breakdown").font = sub_font
    row += 1
    mat_headers = ["Material", "Count", "Volume/Material", "Total Volume", "Used Volume", "Waste", "Utilization"]
    for j, h in enumerate(mat_headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = header_font; c.fill = header_fill; c.border = thin_border
    row += 1
    pattern_map = {p.pattern_id: p for p in result.patterns}
    for mat_name, material in materials.items():
        used_vol = sum(
            count * pattern_map[pid].used_volume
            for pid, count in s.pattern_usage.items()
            if pid in pattern_map and pattern_map[pid].material_name == mat_name
        )
        total_vol = material.volume * material.count
        waste = total_vol - used_vol
        util = used_vol / total_vol if total_vol > 0 else 0.0
        vals = [mat_name, material.count, material.volume, total_vol, used_vol, waste, f"{util * 100:.2f}%"]
        for j, v in enumerate(vals, 1):
            ws.cell(row=row, column=j, value=v).border = thin_border
        row += 1
    row += 1

    # Pattern usage
    ws.cell(row=row, column=1, value="Pattern Usage Details").font = sub_font
    row += 1
    pat_headers = ["Pattern ID", "Material", "Times Used", "Pieces", "Used Volume", "Utilization"]
    for j, h in enumerate(pat_headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = header_font; c.fill = header_fill; c.border = thin_border
    row += 1
    for pid, count in sorted(s.pattern_usage.items()):
        pat = pattern_map.get(pid)
        mat_name = pat.material_name if pat else "?"
        num_pieces = pat.num_pieces if pat else 0
        used_vol = pat.used_volume if pat else 0
        mat = materials.get(mat_name)
        util = used_vol / mat.volume if mat and mat.volume > 0 else 0.0
        vals = [pid, mat_name, count, num_pieces, used_vol, f"{util * 100:.2f}%"]
        for j, v in enumerate(vals, 1):
            ws.cell(row=row, column=j, value=v).border = thin_border
        row += 1

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 22


def fill_result_xlsx(
    template_path: str | Path,
    output_path: str | Path,
    result: ExperimentResult,
    materials: dict[str, Material],
    pieces: dict[str, Piece],
    logger: logging.Logger | None = None,
    timing: dict[str, float] | None = None,
) -> tuple[bool, dict[str, Any]]:
    """Fill a result Excel template with solution data.

    Returns (success, report).
    """
    template_path = Path(template_path)
    output_path = Path(output_path)

    report: dict[str, Any] = {
        "template": str(template_path),
        "output": str(output_path),
        "timestamp": datetime.now().isoformat(),
        "sheets_processed": [],
        "grey_cells_found": 0,
        "grey_cells_filled": 0,
        "grey_cells_skipped": 0,
        "computed_sheet_added": False,
        "errors": [],
        "warnings": [],
    }

    if not template_path.exists():
        msg = f"Template not found: {template_path}"
        if logger: logger.error(msg)
        report["errors"].append(msg)
        return False, report

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(template_path, output_path)

    try:
        wb = openpyxl.load_workbook(output_path)
    except Exception as e:
        msg = f"Failed to open workbook: {e}"
        if logger: logger.error(msg)
        report["errors"].append(msg)
        return False, report

    report["workbook_sheets"] = wb.sheetnames

    # Process each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        grey_cells = _get_grey_cells(ws)
        report["grey_cells_found"] += len(grey_cells)

        # Enrich with context
        for gc in grey_cells:
            gc["adjacent_label"] = _find_adjacent_label(ws, gc["row"], gc["col"])
            gc["column_header"] = _get_column_header(ws, gc["col"])

        filled_count = 0
        skipped_cells: list[str] = []
        for gc in grey_cells:
            cell = ws.cell(row=gc["row"], column=gc["col"])
            filled_value = _determine_cell_value(gc, result, materials, pieces, ws)
            if filled_value is not None:
                cell.value = filled_value
                filled_count += 1
            else:
                skipped_cells.append(gc["coordinate"])

        report["grey_cells_filled"] += filled_count
        report["grey_cells_skipped"] += len(skipped_cells)

        if skipped_cells:
            report["warnings"].append(
                f"Sheet '{sheet_name}': {len(skipped_cells)} grey cells could not be auto-filled: {skipped_cells[:20]}"
            )

        report["sheets_processed"].append({
            "name": sheet_name,
            "grey_cells_found": len(grey_cells),
            "grey_cells_filled": filled_count,
            "grey_cells_skipped": len(skipped_cells),
        })

        if logger:
            logger.info(f"  Sheet '{sheet_name}': {len(grey_cells)} grey, {filled_count} filled, {len(skipped_cells)} skipped")

    # Always add computed_solution sheet
    _add_computed_solution_sheet(wb, result, materials, pieces, timing)
    report["computed_sheet_added"] = True

    try:
        wb.save(output_path)
        if logger: logger.info(f"Saved: {output_path}")
    except Exception as e:
        msg = f"Failed to save: {e}"
        if logger: logger.error(msg)
        report["errors"].append(msg)
        return False, report

    return True, report


def make_backup(template_path: str | Path, backup_dir: str | Path) -> Path:
    """Create a timestamped backup of a template file."""
    template_path = Path(template_path)
    backup_dir = Path(backup_dir)
    backup_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"{template_path.stem}_backup_{ts}{template_path.suffix}"
    shutil.copy(template_path, backup_path)
    return backup_path


def save_template_mapping(
    template_path: str | Path,
    output_path: str | Path,
    logger: logging.Logger | None = None,
) -> dict[str, Any]:
    """Analyze template structure and save grey cell mapping to JSON."""
    template_path = Path(template_path)
    if not template_path.exists():
        return {"error": f"Template not found: {template_path}"}

    wb = openpyxl.load_workbook(template_path)
    sheet_data: dict[str, Any] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        grey_cells = _get_grey_cells(ws)
        for gc in grey_cells:
            gc["adjacent_label"] = _find_adjacent_label(ws, gc["row"], gc["col"])
            gc["column_header"] = _get_column_header(ws, gc["col"])
        sheet_data[sheet_name] = {
            "grey_cell_count": len(grey_cells),
            "grey_cells": grey_cells,
        }

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump({"template": str(template_path), "sheets": sheet_data}, f, indent=2, ensure_ascii=False)

    if logger:
        logger.info(f"Template mapping saved: {output_path}")

    return {"template": str(template_path), "sheets": sheet_data}


def generate_validation_report(
    filled_path: str | Path,
    result: ExperimentResult,
    materials: dict[str, Material],
    pieces: dict[str, Piece],
    output_path: str | Path,
    logger: logging.Logger | None = None,
) -> dict[str, Any]:
    """Read back filled Excel and validate against JSON results."""
    filled_path = Path(filled_path)
    output_path = Path(output_path)

    report: dict[str, Any] = {
        "file": str(filled_path),
        "exists": filled_path.exists(),
        "timestamp": datetime.now().isoformat(),
        "sheet_names": [],
        "has_computed_solution": False,
        "grey_cells_found": 0,
        "grey_cells_filled": 0,
        "metrics_match_json": None,
        "warnings": [],
        "errors": [],
    }

    if not filled_path.exists():
        report["errors"].append("Filled file does not exist")
        return report

    try:
        wb = openpyxl.load_workbook(filled_path)
        report["sheet_names"] = wb.sheetnames
        report["has_computed_solution"] = "computed_solution" in wb.sheetnames

        # Count grey cells that were filled
        for sheet_name in wb.sheetnames:
            if sheet_name == "computed_solution":
                continue
            ws = wb[sheet_name]
            grey_cells = _get_grey_cells(ws)
            filled_count = sum(1 for gc in grey_cells if gc["value"] is not None and gc["value"] != "")
            report["grey_cells_found"] += len(grey_cells)
            report["grey_cells_filled"] += filled_count

        # Cross-check computed_solution sheet against JSON
        if "computed_solution" in wb.sheetnames:
            ws = wb["computed_solution"]
            # Check a few key values
            utilization_cell = _find_cell_by_label(ws, "Material Utilization")
            if utilization_cell:
                try:
                    excel_util = float(str(utilization_cell).replace("%", "")) / 100.0
                    json_util = result.solution.material_utilization
                    if abs(excel_util - json_util) < 0.001:
                        report["metrics_match_json"] = True
                    else:
                        report["metrics_match_json"] = False
                        report["warnings"].append(f"Utilization mismatch: Excel={excel_util}, JSON={json_util}")
                except (ValueError, TypeError):
                    report["warnings"].append("Could not parse utilization from computed_solution sheet")

    except Exception as e:
        report["errors"].append(f"Validation error: {e}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, ensure_ascii=False)

    if logger:
        status = "PASS" if report["metrics_match_json"] and not report["errors"] else "WARN"
        logger.info(f"Excel validation: {status}")

    return report


def _find_cell_by_label(ws, label: str, max_search: int = 50) -> Any:
    """Find cell value to the right of a label in the first columns."""
    for r in range(1, min(max_search, ws.max_row + 1)):
        for c in range(1, 3):
            val = ws.cell(row=r, column=c).value
            if val and str(val).strip().lower() == label.lower():
                return ws.cell(row=r, column=c + 1).value
    return None
