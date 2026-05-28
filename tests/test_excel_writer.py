"""Tests for Excel writer module."""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from openpyxl.styles import PatternFill

from src.cutting3d.excel_writer import (
    _is_grey_fill,
    _get_grey_cells,
    _find_adjacent_label,
    _infer_template_structure,
    save_template_mapping,
)
from src.cutting3d.utils import ensure_dir


def test_is_grey_fill():
    """Test grey fill detection."""
    grey_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    assert _is_grey_fill(grey_fill) is True

    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    assert _is_grey_fill(white_fill) is False

    no_fill = PatternFill()
    assert _is_grey_fill(no_fill) is False


def test_infer_template_structure():
    """Test structure inference on a simple mock sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active

    # Create labels
    ws.cell(row=1, column=1, value="Material")
    ws.cell(row=1, column=2, value="Utilization")
    ws.cell(row=2, column=1, value="L01")

    # Add grey cell at (2, 2)
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=2, column=2).fill = grey_fill

    structure = _infer_template_structure(ws)
    assert len(structure.get("grey_cells", [])) >= 1


def test_save_template_mapping():
    """Test saving template mapping to JSON."""
    # Create a temporary template
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=1, column=1).fill = grey_fill
    ws.cell(row=1, column=1, value="Test")

    tmp_dir = ensure_dir("outputs/tests")
    tmp_path = tmp_dir / "test_template.xlsx"
    wb.save(tmp_path)

    output_path = tmp_dir / "test_mapping.json"
    mapping = save_template_mapping(tmp_path, output_path)
    assert "sheets" in mapping
    assert mapping["sheets"]["Sheet1"]["grey_cell_count"] >= 1


if __name__ == "__main__":
    test_is_grey_fill()
    test_infer_template_structure()
    test_save_template_mapping()
    print("All Excel writer tests passed!")
